#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/10/31 9:37 AM
# @Author  : Ganodermaking

from openpyxl import load_workbook


class Excel:
    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[sheet_name]

    def write(self, row, actual_code, actual_msg):
        self.ws.cell(row=row, column=6).value = actual_code
        self.ws.cell(row=row, column=7).value = actual_msg
        self.wb.save(self.filename)

    def read(self):
        new_list = []
        for row in list(self.ws.rows)[1:]:
            i = 0
            new_dict = {}
            for col in row:
                if i == 1:
                    new_dict['url'] = col.value
                elif i == 2:
                    new_dict['data'] = col.value
                elif i == 3:
                    new_dict['code'] = int(col.value)
                i = i + 1
            new_list.append(new_dict)
        return new_list
